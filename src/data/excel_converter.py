"""
Excel to XML converter for family tree data.
Handles conversion of Excel sheets to XML character files with change detection.
"""
import os
from tkinter import Tk, filedialog
import xml.etree.ElementTree as ET
import pandas as pd
import hashlib
from typing import Tuple, Optional, Dict, Any
import openpyxl
import tkinter as tk
import tkinter.messagebox as messagebox
import traceback
from ..core.path_manager import path_manager

def select_excel_file_and_sheet() -> Tuple[Optional[str], Optional[str]]:
    """
    Show file dialog to select Excel file and get sheet name.
    
    Returns:
        Tuple[str, str]: Excel file path and sheet name, or (None, None) if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    
    if not file_path:
        return None, None
    
    # Load workbook to get sheet names
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        # Create dialog for sheet selection
        dialog = tk.Toplevel(root)
        dialog.title("Select Sheet")
        dialog.geometry("300x150")
        
        selected_sheet = tk.StringVar(value=sheet_names[0])
        
        tk.Label(dialog, text="Choose a worksheet:").pack(pady=10)
        
        for sheet in sheet_names:
            tk.Radiobutton(dialog, text=sheet, variable=selected_sheet, value=sheet).pack()
        
        result = [None]  # Use list to store result from callback
        
        def on_ok():
            result[0] = selected_sheet.get()
            dialog.destroy()
        
        def on_cancel():
            dialog.destroy()
        
        tk.Button(dialog, text="OK", command=on_ok).pack(side=tk.LEFT, padx=20, pady=10)
        tk.Button(dialog, text="Cancel", command=on_cancel).pack(side=tk.RIGHT, padx=20, pady=10)
        
        # Center the dialog
        dialog.transient(root)
        dialog.grab_set()
        root.wait_window(dialog)
        
        return file_path, result[0]
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
        return None, None
    finally:
        root.destroy()

def _generate_xml_string_for_character(character_dict: Dict[str, Any]) -> Optional[str]:
    """
    Generate the XML string for a single character from its dictionary.
    
    Args:
        character_dict: Dictionary containing character data
        
    Returns:
        Optional[str]: XML string representation or None if required fields are missing
    """
    if "ID" not in character_dict or "Name" not in character_dict:
        print(f"Warning: Character dictionary missing 'ID' or 'Name': {character_dict}")
        return None

    root = ET.Element("character", id=str(character_dict["ID"]))

    for col_name, value in character_dict.items():
        if col_name == "ID":
            continue

        xml_tag_name = col_name.replace(" ", "_")
        element = ET.SubElement(root, xml_tag_name)
        if pd.notna(value):
            element.text = str(value)
        else:
            element.text = ""

    return ET.tostring(root, encoding='utf-8', xml_declaration=True).decode('utf-8')

def _sanitize_xml_tag(tag: str) -> str:
    """
    Convert a string into a valid XML tag name.
    Replaces spaces with underscores and removes any invalid characters.
    """
    # Replace spaces with underscores
    tag = tag.replace(' ', '_')
    # Remove any other invalid characters
    tag = ''.join(c for c in tag if c.isalnum() or c == '_')
    # Convert to lowercase for consistency
    return tag.lower()

def create_xml_from_excel_sheet(excel_file_path: str, sheet_name: str, 
                               output_dir: Optional[str] = None) -> str:
    """
    Convert Excel sheet to XML files for family tree processing.
    
    Args:
        excel_file_path: Path to the Excel file
        sheet_name: Name of the sheet to process
        output_dir: Optional output directory (uses path_manager if not provided)
        
    Returns:
        str: Path to the directory containing XML files
    """
    try:
        # Use centralized path management
        if not output_dir:
            excel_base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
            output_dir = path_manager.get_xml_data_dir(excel_base_name, sheet_name)
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        print(f"DEBUG: Converting Excel sheet '{sheet_name}' to XML")
        print(f"DEBUG: Output directory: {output_dir}")
        
        # Read Excel file
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        print(f"DEBUG: Read {len(df)} rows from Excel sheet")
        
        # Group by character (assuming each row is a character)
        characters = {}
        
        for index, row in df.iterrows():
            # Create character ID (use index if no ID column)
            char_id = row.get('id', f"char_{index}")
            if pd.isna(char_id):
                char_id = f"char_{index}"
            
            # Extract character data
            char_data = {}
            for col in df.columns:
                value = row[col]
                if pd.notna(value):
                    char_data[col.lower().replace(' ', '_')] = str(value)
            
            characters[str(char_id)] = char_data
        
        # Create XML files for each character
        for char_id, char_data in characters.items():
            xml_file_path = os.path.join(output_dir, f"{char_id}.xml")
            create_character_xml(char_data, xml_file_path)
        
        print(f"DEBUG: Created {len(characters)} XML files in {output_dir}")
        return output_dir
        
    except Exception as e:
        print(f"ERROR: Failed to convert Excel to XML: {str(e)}")
        raise e

def create_character_xml(char_data: Dict[str, Any], output_file: str):
    """
    Create XML file for a single character.
    
    Args:
        char_data: Dictionary containing character data
        output_file: Path to output XML file
    """
    try:
        # Create root element
        root = ET.Element("character")
        
        # Add character data as elements
        for key, value in char_data.items():
            element = ET.SubElement(root, key)
            element.text = str(value)
        
        # Create tree and write to file
        tree = ET.ElementTree(root)
        tree.write(output_file, encoding='utf-8', xml_declaration=True)
        
    except Exception as e:
        print(f"ERROR: Failed to create XML for character: {str(e)}")
        raise e

def create_xml_from_excel_sheet(excel_file_path: str, sheet_name: str, output_xml_dir: str) -> str:
    """
    Create XML files from Excel sheet data.
    
    Args:
        excel_file_path: Path to the Excel file
        sheet_name: Name of the sheet to process
        output_xml_dir: Directory to save XML files
    
    Returns:
        str: Path to the directory containing generated XML files
    """
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_xml_dir, exist_ok=True)
        
        print(f"\nDEBUG: Processing Excel file: {excel_file_path}")
        print(f"DEBUG: Sheet name: {sheet_name}")
        print(f"DEBUG: Output directory: {output_xml_dir}")
        
        # Load Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        
        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = workbook[sheet_name]
        print(f"DEBUG: Worksheet loaded, dimensions: {worksheet.dimensions}")
        
        # Get headers from first row and sanitize them
        headers = [cell.value for cell in worksheet[1] if cell.value]
        print(f"DEBUG: Headers found: {headers}")
        
        # Track which XML files to delete
        existing_files = set()
        for file in os.listdir(output_xml_dir):
            if file.endswith('.xml'):
                existing_files.add(os.path.join(output_xml_dir, file))
        
        # Process each row
        files_created = set()
        for row in worksheet.iter_rows(min_row=2):  # Skip header row
            # Get values and create dict
            row_data = {headers[i]: cell.value for i, cell in enumerate(row) if cell.value is not None and i < len(headers)}
            
            if not row_data:
                continue  # Skip empty rows
            
            # Extract ID
            char_id = row_data.get('ID')
            if not char_id:
                print(f"DEBUG: Skipping row without ID: {row_data}")
                continue
            
            xml_file = os.path.join(output_xml_dir, f"{char_id}.xml")
            files_created.add(xml_file)
            
            # Create XML structure
            root = ET.Element("character")
            
            # Add ID first
            id_elem = ET.SubElement(root, "id")
            id_elem.text = str(char_id)
            
            # Add name elements
            if 'Name' in row_data:
                name_elem = ET.SubElement(root, "name")
                name_elem.text = str(row_data['Name'])
            
            if 'Middle Name' in row_data and row_data['Middle Name']:
                middle_name_elem = ET.SubElement(root, "middle_name")
                middle_name_elem.text = str(row_data['Middle Name'])
            
            # Add other standard fields
            field_mappings = {
                'Birthday': 'birthday',
                'Birth Place': 'birth_place',
                'Died': 'died',
                'Date of Death': 'date_of_death',
                'Place of Burial': 'place_of_burial',
                'Marital Status': 'marital_status',
                'Marriage Date': 'marriage_date',
                'Place of Marriage': 'place_of_marriage',
                'Spouse': 'spouse',
                'Spouse ID': 'spouse_id',
                'Father': 'father',
                'Father ID': 'father_id',
                'Mother': 'mother',
                'Mother ID': 'mother_id'
            }
            
            for excel_field, xml_field in field_mappings.items():
                if excel_field in row_data and row_data[excel_field] is not None:
                    elem = ET.SubElement(root, xml_field)
                    elem.text = str(row_data[excel_field])
            
            # Add any remaining fields not in standard mappings
            for field, value in row_data.items():
                if (field not in ['ID', 'Name', 'Middle Name'] and 
                    field not in field_mappings and 
                    value is not None):
                    elem = ET.SubElement(root, _sanitize_xml_tag(field))
                    elem.text = str(value)
            
            # Write XML file
            tree = ET.ElementTree(root)
            tree.write(xml_file, encoding='utf-8', xml_declaration=True)
            print(f"DEBUG: Created XML file: {xml_file}")
        
        # Delete any old files that weren't recreated
        files_to_delete = existing_files - files_created
        for file in files_to_delete:
            try:
                os.remove(file)
                print(f"DEBUG: Deleted old XML file: {file}")
            except OSError as e:
                print(f"Warning: Could not delete old XML file {file}: {e}")
        
        workbook.close()
        return output_xml_dir
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        if hasattr(e, '__traceback__'):
            print(traceback.format_exc())
        raise
    finally:
        if 'workbook' in locals():
            workbook.close() 